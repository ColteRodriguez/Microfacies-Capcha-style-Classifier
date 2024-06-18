import openpyxl

# Methods for spreadsheet manipulation
def locate_sample(spreadsheet_path, name_query, sheet_num):
    workbook = openpyxl.load_workbook(spreadsheet_path)
    sheet = workbook[workbook.sheetnames[sheet_num]]
    # Only search the first row becasue the sample name cant be anywhere else
    for row in sheet.iter_rows():
        for cell in row:
            if str(cell.value) == name_query:
                return cell.row, cell.column
            
    raise Exception(f"Query for {name_query} at location {spreadsheet_path} returned Null")
    return
        
def get_cell(row, col, spreadsheet_path, sheet_num):
    workbook = openpyxl.load_workbook(spreadsheet_path)
    sheet = workbook[workbook.sheetnames[sheet_num]]
    
    val = sheet.cell(row=row, column=col).value
    workbook.save(spreadsheet_path)
    return val

def change_cell(row, col, spreadsheet_path, val, sheet_num):
    workbook = openpyxl.load_workbook(spreadsheet_path)
    sheet = workbook[workbook.sheetnames[sheet_num]]
    
    sheet.cell(row=row, column=col, value=val)
    workbook.save(spreadsheet_path)
    return val

def update_percent(Pcurr, Tcurr, segment_area, rating, isSubject):
    # Div by 0 error should not occur because total area is always updated 
    # Update the rating
    if rating is not None:
        return ((Pcurr * Tcurr) + (rating * segment_area)) / (Tcurr + segment_area)
    # Update the constituent or feature
    elif isSubject == True:
        return ((Pcurr * Tcurr) + (segment_area)) / (Tcurr + segment_area)
    # Update other constituents whose areas are not to be updated
    else:
        return (Pcurr * Tcurr) / (Tcurr + segment_area)
    
    
def add_new_sample(path, name):
    def find_first_empty_row_header(path):
        # Load the workbook and select the specified sheet
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[workbook.sheetnames[0]]

        # Start from row 2 since row 1 is the header
        for row in range(2, sheet.max_row + 2):
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value is None or cell_value == "":
                return row
        return None  # Return None if no empty cell is found

    row = find_first_empty_row_header(path)
    print(row)
    
    # Add in the name and fill in 0s
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[workbook.sheetnames[0]]
    sheet.cell(row=row, column=1, value=name)
    for i in range(2,sheet.max_column + 1):
        sheet.cell(row=row, column=i, value=0)
        
    # Add in the name and fill in 0s
    sheet = workbook[workbook.sheetnames[1]]
    sheet.cell(row=row, column=1, value=name)
    for i in range(2,sheet.max_column + 1):
        sheet.cell(row=row, column=i, value=0)
    workbook.save(path)
    